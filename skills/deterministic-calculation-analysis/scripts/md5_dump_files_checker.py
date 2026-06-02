#!/usr/bin/env python3
"""检查msprobe数据，确定分析级别（L1或mix）。"""

import argparse
import csv
import os
import sqlite3


def find_first_dump_file(root_path):
    for dirpath, _, filenames in os.walk(root_path):
        for f in filenames:
            if f == 'dump.json':
                return os.path.join(dirpath, f)
    return None


def detect_path_type(root_path):
    """检测路径类型: 'db', 'csv_xlsx', 或 None"""
    names = [root_path] if os.path.isfile(root_path) else os.listdir(root_path)
    for name in names:
        if name.endswith('.vis.db'):
            return 'db'
        if name.endswith(('.xlsx', '.csv')):
            return 'csv_xlsx'
    return None


def _first_file(root_path, exts):
    if os.path.isfile(root_path):
        return root_path
    for name in sorted(os.listdir(root_path)):
        if name.endswith(exts):
            return os.path.join(root_path, name)
    return None


def _check_l0_names(names_iter, max_scan=100):
    """扫描前max_scan个name，全为Module./Cell.前缀则返回True。"""
    count = 0
    for i, name in enumerate(names_iter):
        if i >= max_scan:
            break
        if name:
            count += 1
            if not name.startswith(('Module.', 'Cell.')):
                return False
    return count > 0


def validate_csv_xlsx(root_path):
    """校验csv/xlsx文件头，并检测是否为L0（仅Module级数据）。"""
    first = _first_file(root_path, ('.csv', '.xlsx'))
    if not first:
        return "未找到 .csv 或 .xlsx 文件"
    try:
        if first.endswith('.csv'):
            with open(first, encoding='utf-8') as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                if _check_l0_names((row.get('NPU Name', '') for row in reader)):
                    return "当前数据仅包含Module级信息，没有API级数据，无法分析确定性问题。"
        else:
            import openpyxl
            wb = openpyxl.load_workbook(first, read_only=True)
            ws = wb.active
            fieldnames = [str(c.value) if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if _check_l0_names(str(row[0]) if row[0] is not None else '' for row in ws.iter_rows(min_row=2, values_only=True)):
                wb.close()
                return "当前数据仅包含Module级信息，没有API级数据，无法分析确定性问题。"
            wb.close()
    except Exception as e:
        return f"读取文件失败: {first}\n  {e}"
    missing = [c for c in ('NPU MD5', 'BENCH MD5') if c not in fieldnames]
    if missing:
        return f"缺少比对字段: {', '.join(missing)}，没有包含tensor的CRC-32校验值，无法分析确定性问题。"
    return None


def validate_db(root_path):
    """校验db的tb_config表，并检测是否为L0（无API节点）。"""
    first = _first_file(root_path, ('.vis.db',))
    if not first:
        return "未找到 .vis.db 文件"
    try:
        conn = sqlite3.connect(f"file:{first}?mode=ro", uri=True)
        task_values = {row[0] for row in conn.execute("SELECT task FROM tb_config")}
        has_api = any(row[0] == '1' for row in conn.execute("SELECT DISTINCT node_type FROM tb_nodes"))
        conn.close()
    except Exception as e:
        return f"读取db文件失败: {first}\n  {e}"
    if 'md5' not in task_values:
        return "tb_config 表的 task 字段不是 md5，没有包含tensor的CRC-32校验值，无法分析确定性问题。"
    if not has_api:
        return "当前数据仅包含Module级信息，没有API级数据，无法分析确定性问题。"
    return None


def check_dump_file(filepath, label):
    """检查dump.json前100行，返回level值或抛出异常。"""
    if not filepath:
        raise RuntimeError(f"({label}) 未找到 dump.json 文件")
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = [f.readline() for _ in range(100)]
    except Exception as e:
        raise RuntimeError(f"({label}) 读取文件失败: {filepath}\n  {e}")
    content = ''.join(lines)

    if '"md5":' not in content:
        raise RuntimeError(f"({label}) 当前dump数据没有包含tensor的CRC-32校验值，无法分析确定性问题。\n  文件: {filepath}")
    level = None
    for line in lines:
        stripped = line.strip().rstrip(',')
        if stripped.startswith('"level"'):
            level = stripped.split(':', 1)[-1].strip().strip('"')
            break
    if level is None:
        raise RuntimeError(f"({label}) dump.json 中未找到 level 字段。\n  文件: {filepath}")
    if level not in ('L1', 'mix'):
        raise RuntimeError(f'({label}) dump数据的level="{level}"，需要为"L1"或"mix"。\n  文件: {filepath}')
    return level


def main():
    parser = argparse.ArgumentParser(description='检查msprobe数据，确定分析级别（L1或mix）。')
    parser.add_argument('target', help='dump target路径，或 db/csv/xlsx 路径')
    parser.add_argument('golden', nargs='?', help='dump golden路径（db/csv/xlsx 路径不需要）')
    args = parser.parse_args()

    if args.golden is None:
        if not os.path.exists(args.target):
            parser.exit(1, f"错误: 路径不存在: {args.target}\n")
        ptype = detect_path_type(args.target)
        if ptype == 'db':
            err = validate_db(args.target)
        elif ptype == 'csv_xlsx':
            err = validate_csv_xlsx(args.target)
        else:
            parser.exit(1, f"错误: 未找到 .vis.db 或 .csv/.xlsx 文件: {args.target}\n")
        if err:
            parser.exit(1, f"错误: {ptype}文件校验不通过。\n  {err}\n")
        level = 'mix' if ptype == 'db' else 'L1'
        print(f'level="{level}"')
        return

    target_path, golden_path = args.target, args.golden
    for p in (target_path, golden_path):
        if not os.path.exists(p):
            parser.exit(1, f"错误: 路径不存在: {p}\n")

    target_file = find_first_dump_file(target_path)
    golden_file = find_first_dump_file(golden_path)

    all_pass = True
    levels = {}
    for filepath, label in [(target_file, 'target'), (golden_file, 'golden')]:
        try:
            levels[label] = check_dump_file(filepath, label)
        except RuntimeError as e:
            print(e)
            all_pass = False

    if all_pass and len(levels) == 2 and levels['target'] != levels['golden']:
        print(f"target和golden的level不一致: target=\"{levels['target']}\", golden=\"{levels['golden']}\"")
        all_pass = False

    if all_pass:
        print(f"level=\"{levels['target']}\"")
    else:
        parser.exit(1)


if __name__ == "__main__":
    main()
