#!/usr/bin/env python3
"""分析msprobe L1级别比对文件（xlsx/csv），寻找首个输入一致输出不一致的API。"""

import argparse
import csv
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

from msprobe_utils import build_excluded_rules, is_key_excluded, print_kv_table, is_api_excluded


def parse_api_name(npu_name):
    for io_type in ('input', 'output'):
        m = re.match(r'^(.*)\.' + io_type + r'\.(.+)$', npu_name)
        if m:
            return m.group(1), io_type, m.group(2)
    return None, None, None


def read_csv_rows(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def read_xlsx_rows(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row_cells):
            if i < len(header):
                row_dict[header[i]] = str(val) if val is not None else ''
        rows.append(row_dict)
    wb.close()
    return rows


def check_api(api_name, inputs, outputs, rules):
    """检查单个API: 输入md5是否都一致, 输出是否有不一致"""
    input_match = True
    for idx, npu_md5, bench_md5 in inputs:
        if is_key_excluded(api_name, 'input', idx, rules):
            continue
        if npu_md5 != bench_md5:
            input_match = False
            break

    output_diffs = []
    for idx, npu_md5, bench_md5 in outputs:
        if is_key_excluded(api_name, 'output', idx, rules):
            continue
        if npu_md5 != bench_md5:
            output_diffs.append((idx, npu_md5, bench_md5))

    output_match = len(output_diffs) == 0
    return {
        'input_match': input_match,
        'output_match': output_match,
        'output_diffs': output_diffs if not output_match else [],
        'input_keys': [(idx, npu_md5) for idx, npu_md5, _ in inputs
                      if not is_key_excluded(api_name, 'input', idx, rules)],
    }


def analyze_file(filepath, excluded_apis=None):
    """分析单个L1比对文件，返回该文件的rank及分析结果"""
    rules = build_excluded_rules()
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        rows = read_csv_rows(filepath)
    elif ext == '.xlsx':
        rows = read_xlsx_rows(filepath)
    else:
        return {'rank': 0, 'found': False, 'reason': f'不支持的文件格式: {ext}'}

    if not rows:
        return {'rank': 0, 'found': False, 'reason': '文件为空'}

    rank = 0
    m = re.search(r'rank(\d+)', os.path.basename(filepath))
    if m:
        rank = int(m.group(1))

    # ── 第一遍扫描：按API分组，检查每个API的状态 ──
    api_status_list = []
    current_api = None
    current_inputs = []
    current_outputs = []
    for row in rows:
        npu_name = row.get('NPU Name', '')
        if not npu_name:
            continue
        api_name, io_type, idx = parse_api_name(npu_name)
        if api_name is None:
            continue
        if current_api is not None and api_name != current_api:
            api_status_list.append((current_api, check_api(current_api, current_inputs, current_outputs, rules)))
            current_inputs = []
            current_outputs = []
        current_api = api_name
        npu_md5 = row.get('NPU MD5', '').strip()
        bench_md5 = row.get('BENCH MD5', '').strip()
        if io_type == 'input':
            current_inputs.append((idx, npu_md5, bench_md5))
        else:
            current_outputs.append((idx, npu_md5, bench_md5))

    if current_api is not None:
        api_status_list.append((current_api, check_api(current_api, current_inputs, current_outputs, rules)))

    # ── 第二遍：查找首个输入一致输出不一致的API ──
    for api_name, status in api_status_list:
        if is_api_excluded(api_name, excluded_apis):
            continue
        if status['input_match'] and not status['output_match']:
            return {
                'rank': rank, 'found': True, 'api': api_name,
                'input_keys': status['input_keys'], 'output_diffs': status['output_diffs'],
            }

    last_good = (-1, None)
    first_bad = (-1, None)
    for i, (api_name, status) in enumerate(api_status_list):
        if is_api_excluded(api_name, excluded_apis):
            continue
        if status['input_match'] and status['output_match']:
            last_good = (i, api_name)
        elif not status['input_match']:
            first_bad = (i, api_name)
            break

    result = dict(rank=rank, found=False, reason='未找到输入一致输出不一致的API')
    if last_good[1] or first_bad[1]:
        t = dict(last_good_api=last_good[1], first_bad_api=first_bad[1],
                 last_good_idx=last_good[0], first_bad_idx=first_bad[0])
        if last_good[0] >= 0 and first_bad[0] >= 0 and first_bad[0] - last_good[0] > 1:
            t['between_apis'] = [api_status_list[j][0] for j in range(last_good[0] + 1, first_bad[0])]
        result['transition'] = t
    return result


def analyze_path(path, excluded_apis=None):
    if not os.path.exists(path):
        print(f"错误: 路径不存在: {path}")
        return

    if os.path.isfile(path):
        files = [path]
    else:
        files = sorted(os.path.join(path, f) for f in os.listdir(path) if f.endswith(('.xlsx', '.csv')))

    if not files:
        print(f"未找到 xlsx 或 csv 文件: {path}")
        return

    print(f"共发现 {len(files)} 个文件")
    print("=" * 80)

    all_results = []
    with ThreadPoolExecutor(max_workers=min(len(files), 32)) as executor:
        future_to_file = {executor.submit(analyze_file, fp, excluded_apis): fp for fp in files}
        for future in as_completed(future_to_file):
            fp = future_to_file[future]
            try:
                all_results.append(future.result())
            except Exception as e:
                rank = 0
                m = re.search(r'rank(\d+)', os.path.basename(fp))
                if m:
                    rank = int(m.group(1))
                all_results.append({'rank': rank, 'found': False, 'reason': f"分析异常: {e}"})

    all_results.sort(key=lambda r: r['rank'])

    # ── 每个rank一张双列表 ──
    for r in all_results:
        kv_rows = []
        if r['found']:
            kv_rows.append(('首个问题API', r['api']))
            items = ['Input MD5 (全部一致):'] + [f'  {idx}: {md5}' for idx, md5 in r['input_keys']]
            items += ['Output MD5 (不一致):'] + [f'  {idx}: NPU={npu_md5} vs Bench={bench_md5}' for idx, npu_md5, bench_md5 in r['output_diffs']]
            kv_rows.append(('API分析依据', '\n'.join(items)))
        else:
            kv_rows.append(('首个问题API', '无'))
            t = r.get('transition')
            if t:
                lines = [r.get('reason', '')]
                if t.get('last_good_api'):
                    lines.extend(['', '最后一个完全正常的API (输入匹配+输出匹配):', f"  {t['last_good_api']}"])
                if t.get('first_bad_api'):
                    lines.extend(['第一个输入不匹配的API:', f"  {t['first_bad_api']}"])
                if t.get('between_apis'):
                    lines.append('两者之间的API (可能有异常或被漏采):')
                    for ba in t['between_apis']:
                        lines.append(f"  {ba}")
                lines.extend(['', '可能原因:'])
                if t.get('last_good_api'):
                    lines.append(f'  1. "{t["last_good_api"]}" 的输出发生改变，但结果集可能被msprobe漏采')
                    lines.append(f'  2. 漏采的API位于 "{t["last_good_api"]}" 和 "{t["first_bad_api"]}" 之间')
                lines.append('  3. 通信API (all_reduce等) 传递了其他rank的diff结果')
                kv_rows.append(('API分析依据', '\n'.join(lines)))
            else:
                kv_rows.append(('API分析依据', r.get('reason', '未找到首个输入一致输出不一致的API')))
        print_kv_table(r['rank'], kv_rows)

    found_apis = [r for r in all_results if r['found']]
    if found_apis and not excluded_apis:
        print("\n提示: 如果不认为上述API是问题根因，可输入API名称排除后重新分析。支持前缀匹配，多个API以空格分隔。")


def main():
    parser = argparse.ArgumentParser(description='分析msprobe L1级别比对文件，寻找首个输入一致输出不一致的API。')
    parser.add_argument('path', help='比对结果目录或文件（.xlsx/.csv）')
    parser.add_argument('--exclude-api', nargs='+', default=[], metavar='NAME',
                        help='要排除的API名称（支持前缀匹配，多个以空格分隔）')
    args = parser.parse_args()

    excluded = set(args.exclude_api)
    if excluded:
        print(f"排除以下API前缀: {sorted(excluded)}")
    analyze_path(args.path, excluded)


if __name__ == "__main__":
    main()
